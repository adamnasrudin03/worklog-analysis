"""Export laporan ke Excel (.xlsx)."""

from __future__ import annotations

from pathlib import Path

from worklog_lib import (
    ReportData,
    format_day_excel_field,
    format_day_grouping_field,
    format_hours,
)


def export_xlsx(report: ReportData, path: Path, field: str = "summary") -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportError(
            "openpyxl diperlukan untuk export xlsx. Install: pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Tanggal", "Grouping", "Ringkasan"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for day in report.daily_summaries:
        ws.append(
            [
                day.day,
                format_day_grouping_field(day),
                format_day_excel_field(day, field),
            ]
        )

    ws2 = wb.create_sheet("KPI")
    ws2.append(["Metrik", "Nilai"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    rows = [
        ("Periode", report.period_label),
        ("Total jam", format_hours(report.total_hours)),
        ("Entri", len(report.entries)),
        ("DONE", len(report.done_keys)),
    ]
    if report.sp_metrics.total_sp > 0:
        hps = report.sp_metrics.hours_per_sp
        rows.append(
            (
                "Jam/SP",
                f"{hps:.2f}" if hps is not None else "-",
            )
        )
    for label, value in rows:
        ws2.append([label, value])

    wb.save(path)
